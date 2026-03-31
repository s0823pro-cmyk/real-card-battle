# -*- coding: utf-8 -*-
"""大工カード一覧を carpenter_cards.xlsx に出力（openpyxl）。"""
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DECK = ROOT / "src" / "data" / "carpenterDeck.ts"
JOB = ROOT / "src" / "data" / "jobs" / "carpenter.ts"
EXP = ROOT / "src" / "data" / "jobs" / "carpenterExpansion.ts"
OUT = ROOT / "carpenter_cards.xlsx"

HEADERS = [
    "カテゴリ",
    "プール定数名",
    "ID",
    "カード名（日本語）",
    "タイプ",
    "コスト(秒)",
    "効果説明",
    "レアリティ",
    "備考",
]


def extract_block_after_id(text: str, card_id: str) -> str:
    """id: 'card_id' を含むオブジェクトリテラル { ... } をブレース対応で取得（1行／複数行両対応）。"""
    needle = f"id: '{card_id}'"
    i = text.find(needle)
    if i < 0:
        raise ValueError(f"id not found: {card_id}")
    start = text.rfind("{", 0, i)
    if start < 0:
        raise ValueError(f"no opening brace for {card_id}")
    depth = 0
    for k in range(start, len(text)):
        if text[k] == "{":
            depth += 1
        elif text[k] == "}":
            depth -= 1
            if depth == 0:
                return text[start : k + 1]
    raise ValueError(f"unclosed brace for {card_id}")


def pick_str(block: str, key: str) -> str | None:
    m = re.search(rf"{key}:\s*'((?:\\'|[^'])*)'", block)
    return m.group(1).replace("\\'", "'") if m else None


def pick_num(block: str, key: str) -> str | None:
    m = re.search(rf"{key}:\s*([0-9]+(?:\.[0-9]+)?)", block)
    return m.group(1) if m else None


def build_effect_description(block: str) -> str:
    desc = pick_str(block, "description") or ""
    rb = re.search(
        r"reserveBonus:\s*\{[\s\S]*?description:\s*'((?:\\'|[^'])*)'",
        block,
    )
    extra = []
    if rb:
        extra.append(rb.group(1).replace("\\'", "'"))
    if extra:
        return desc + (" / " if desc else "") + " / ".join(extra)
    return desc


def format_cost(block: str) -> str:
    tc = pick_num(block, "timeCost")
    prep = pick_num(block, "preparationTimeCost")
    if not tc:
        return ""
    if prep:
        return f"{tc}（段取り{prep}秒）"
    return tc


def pick_rarity(block: str) -> str:
    r = pick_str(block, "rarity")
    if r == "rare":
        return "rare"
    return ""


def parse_card(text: str, card_id: str) -> dict[str, str]:
    block = extract_block_after_id(text, card_id)
    return {
        "id": card_id,
        "name": pick_str(block, "name") or "",
        "type": pick_str(block, "type") or "",
        "cost": format_cost(block),
        "effect": build_effect_description(block),
        "rarity": pick_rarity(block),
    }


# 行定義: (カテゴリ, プール定数名, ファイルパス, id, 備考オーバーライドまたは None)
ROWS: list[tuple[str, str, Path, str, str | None]] = []

# --- スターター ---
for sid in [
    "hammer_1",
    "hammer_2",
    "hammer_3",
    "hammer_4",
    "saw_guard_1",
    "saw_guard_2",
    "saw_guard_3",
    "build_scaffold",
    "nail_strike",
    "work_clothes",
]:
    ROWS.append(
        (
            "スターター",
            "CARPENTER_STARTER_DECK / buildStarterDeck",
            DECK,
            sid,
            "buildStarterDeck の assignId により実行時は id に _1, _2 … の連番が付く",
        )
    )

# --- 温存ボーナス ---
for sid in ["aged_wood", "sharpened_saw", "reinforced_wall"]:
    note = "CARPENTER_UNCOMMON_POOL_UNFILTERED にもスプレッドで含まれる（reinforced_wall はコモンプールにも別定義あり）" if sid != "reinforced_wall" else None
    if sid == "reinforced_wall":
        note = "carpenter.ts の CARPENTER_COMMON_POOL_UNFILTERED にも同一IDで定義（重複）"
    ROWS.append(("温存ボーナス", "RESERVE_BONUS_CARDS", DECK, sid, note))

# --- デバフ ---
ROWS.append(("デバフ", "ANXIETY_CARD", DECK, "anxiety", None))
ROWS.append(("デバフ", "CURSE_CARD", DECK, "curse", None))

# --- コモン（直書き6枚。reinforced_wall は上と重複定義の2行目） ---
for sid in [
    "power_drill",
    "wood_block",
    "blueprint_draw",
    "sumidashi",
    "quick_hammer",
]:
    ROWS.append(("コモン", "CARPENTER_COMMON_POOL_UNFILTERED", JOB, sid, None))
ROWS.append(
    (
        "コモン",
        "CARPENTER_COMMON_POOL_UNFILTERED",
        JOB,
        "reinforced_wall",
        "carpenterDeck.ts の RESERVE_BONUS_CARDS と同一IDの重複定義",
    )
)

# --- コモン拡張 ---
EXP_COMMON_IDS = [
    "kiso_ashiba",
    "sumi_hosoku",
    "ko_kugiuchi",
    "kakuzai",
    "kari_gakou",
    "suiheiki",
    "kikuzu_harai",
    "shitami_ita",
    "kanazuchi_tap",
    "sumitsubo_makijaku",
    "yojo_tape",
    "kui_uchi",
    "shiguchi_check",
    "kanejaku_ate",
    "daiku_nomi_mejirushi",
    "sagyo_dai",
    "yaneura_kakunin",
    "kugibukuro_seiri",
    "mokufun_harai",
    "taruki_no_shita",
    "sumitsuke_naoshi",
    "genba_haki",
]
for sid in EXP_COMMON_IDS:
    ROWS.append(("コモン拡張", "CARPENTER_EXPANSION_COMMON", EXP, sid, None))

# --- アンコモン（直書き。温存2枚は定義元が RESERVE のため行は作らず備考のみ温存側） ---
for sid in [
    "large_crane",
    "defense_wall",
    "foreman",
    "reinforced_concrete",
    "safety_helmet",
    "iron_wall",
]:
    ROWS.append(("アンコモン", "CARPENTER_UNCOMMON_POOL_UNFILTERED", JOB, sid, None))

# --- アンコモン拡張 ---
EXP_UNCOMMON_IDS = [
    "hashira_tateru",
    "hari_tsugite",
    "daiku_hashigo",
    "kanna_kuzu_tobashi",
    "kanazuchi_hibiki",
    "dodai_katame",
    "toshi_bashira",
    "yane_fuki",
    "genkan_waku",
    "nokogiri_renda",
    "naiso_shitaji",
    "yukabari",
    "kensa_gokaku",
    "measure_sokutei",
    "shiage_kanna",
]
for sid in EXP_UNCOMMON_IDS:
    ROWS.append(("アンコモン拡張", "CARPENTER_EXPANSION_UNCOMMON", EXP, sid, None))

# --- レア ---
for sid in ["mega_nail", "renovation"]:
    ROWS.append(("レア", "CARPENTER_RARE_POOL_UNFILTERED", JOB, sid, None))

# --- レア拡張 ---
EXP_RARE_IDS = [
    "cho_mabashira",
    "zenmen_kaiso",
    "koryo_setsugo",
    "niju_ashiba",
    "tenken_sha",
    "meisho_nomi",
    "ishizue_ichigeki",
]
for sid in EXP_RARE_IDS:
    ROWS.append(("レア拡張", "CARPENTER_EXPANSION_RARE", EXP, sid, None))

# --- 実績レア ---
for sid in ["ridgepole", "temple_carpenter", "master_strike"]:
    ROWS.append(("実績レア", "CARPENTER_ACHIEVEMENT_RARE_CARDS", JOB, sid, None))


def main() -> None:
    deck_txt = DECK.read_text(encoding="utf-8")
    job_txt = JOB.read_text(encoding="utf-8")
    exp_txt = EXP.read_text(encoding="utf-8")

    file_map = {DECK: deck_txt, JOB: job_txt, EXP: exp_txt}

    wb = Workbook()
    ws = wb.active
    ws.title = "大工カード"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for cat, pool, path, cid, note_override in ROWS:
        text = file_map[path]
        try:
            d = parse_card(text, cid)
        except ValueError as e:
            raise SystemExit(f"parse error {cid} in {path}: {e}") from e
        note = note_override if note_override is not None else ""
        ws.append(
            [
                cat,
                pool,
                d["id"],
                d["name"],
                d["type"],
                d["cost"],
                d["effect"],
                d["rarity"],
                note,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 48
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 42

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({ws.max_row - 1} rows)")


if __name__ == "__main__":
    main()
